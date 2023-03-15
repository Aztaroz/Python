import openpyxl
from openpyxl.utils import get_column_letter

def notCover(sheet):
    if sheet.title.upper() != 'COVER':
        return True
    else:
        return False
    
def inchToCent(inch): # แปลงจากนิ้วเป็นเซนติเมตร
	inch = inch * 2.54
	return round(inch, 2)

def marginTop(margin): 
	if sheet.title.upper() != "COVER":
		title = sheet.title
		base = 2.4
		margin = inchToCent(margin.page_margins.top) - base
		if abs(margin) <= 0.5: #0.7 = False
			return True
		else: return print(f'Worksheet : "{title}" Margin Top ต้องเป็น {base} เท่านั้น ค่าของคุณ : {margin + base} ')
	else: return sheet.title
    
def marginLeft(margin):
	if sheet.title.upper() != "COVER":
		title = sheet.title
		base = 1.8
		margin = inchToCent(margin.page_margins.left) - base
		if abs(margin) <= 0.5:
			return True 
		else: return print(f'Worksheet : "{title}" Margin Left ต้องเป็น {base} เท่านั้น ค่าของคุณ : {margin + base} ')
	else: return sheet.title
    
def marginRight(margin):
	if sheet.title.upper() != "COVER":
		title = sheet.title
		base = 1.8
		margin = inchToCent(margin.page_margins.right) - base
		if abs(margin) <= 0.5:
			return True
		else: return print(f'Worksheet : "{title}" Margin Right ต้องเป็น {base} เท่านั้น ค่าของคุณ : {margin + base} ')
	else: return sheet.title
    
def marginBottom(margin):
	if sheet.title.upper() != "COVER":
		title = sheet.title
		base = 1.8
		margin = inchToCent(margin.page_margins.bottom) - base
		if abs(margin) <= 0.5:
			return True
		else: return print(f'Worksheet : "{title}" Margin Bottom ต้องเป็น {base} เท่านั้น ค่าของคุณ : {margin + base} ')
	else: return sheet.title

def marginHeader(margin):
	if sheet.title.upper() != "COVER":
		title = sheet.title
		base = 0.8
		margin = inchToCent(margin.page_margins.header) - base
		if abs(margin) <= 0.5:
			return True
		else: return print(f'Worksheet : "{title}" Margin Header ต้องเป็น {base} เท่านั้น ค่าของคุณ : {margin + base} ')
	else: return sheet.title 
    
def marginFooter(margin):
	if sheet.title.upper() != "COVER":
		title = sheet.title
		base = 0.8
		margin = inchToCent(margin.page_margins.footer) - base
		if abs(margin) <= 0.5:
			return True
		else: return print(f'Worksheet : "{title}" Margin Footer ต้องเป็น {base} เท่านั้น ค่าของคุณ : {margin + base} ')
	else: return sheet.title

def horizontalCenter(sheet):
	if sheet.title.upper() != "COVER":
		title = sheet.title
		sheet = sheet.print_options.horizontalCentered
		if sheet is None:
			return print(f'Worksheet : "{title}" Center on Page ต้องเป็น Horizontally เท่านั้น')
		else: return True
	else: return sheet.title

def headerCheck(sheet): #มี Score Paremeter?
    center = sheet.oddHeader.center
    font = sheet.oddHeader.center.font.startswith('TH SarabunPSK')
    isBold = sheet.oddHeader.center.font.endswith('Bold') or sheet.oddHeader.center.font.endswith('ตัวหนา')
    if center.text.replace(" ","") == "ยอดขายเครื่องสำอางของพนักงานขายไตรมาสแรกของปี2562-2564":
        if font == True:
            if center.size == 22 :
                if isBold == True:
                    return True
                else: print(f'Worksheet : "{sheet.title}"Font ที่ Header ต้องเป็น ตัวหนา(Bold) เท่านั้น')
            else : print(f'Worksheet : "{sheet.title}" ขนาด Font ที่ Header ต้องเป็นขนาด 22 เท่านั้น ขนาด Font ของคุณ {center.size}')
        else: print(f'Worksheet : "{sheet.title}"Font ที่ Header ต้องเป็น TH SarabunPSK เท่านั้น')
    else: print(f'Worksheet : "{sheet.title}" Header Center ต้องเป็น "ยอดขายเครื่องสำอางของพนักงานขายไตรมาสแรกของปี 2562-2564"')

def footerCheck(sheet,score):
    left = sheet.oddFooter.left
    center = sheet.oddFooter.center
    right = sheet.oddFooter.right
    # print(left.font)
    if left.text:
        if left.font != None:
            if left.font.startswith("TH SarabunPSK"):
                if left.size == 18:
                    if left.font.endswith("Bold") or left.font.endswith("ตัวหนา"): score += 1
                    else : print(f'Worksheet : "{sheet.title}" Font ที่ Footer Left Section ต้องเป็น ตัวหนา(Bold) เท่านั้น')
                else: print(f'Worksheet : "{sheet.title}" ขนาด Font ที่ Footer left Section ต้องมีขนาด 18 เท่านั้น ขนาด Font ของคุณ : {left.size}')
            else: print(f'Worksheet : "{sheet.title}" Font ที่ Footer Left Section ต้องเป็น TH SarabunPSK เท่านั้น Font ของคุณ : {left.font}')
        else: print(f'Worksheet : "{sheet.title}" Font ที่ Footer Left Section ต้องเป็น TH SarabunPSK เท่านั้น Font ของคุณ : Calibri')
    else : print(f'Worksheet : "{sheet.title}" Footer Left section กำหนดให้พิมพ์รหัสนักศึกษาของตนเอง')
        
    if center.text:
        if center.font != None:
            if center.font.startswith("TH SarabunPSK"):
                if center.size == 18: 
                    if center.font.endswith("Bold") or center.font.endswith("ตัวหนา"): score += 1
                    else : print(f'Worksheet : "{sheet.title}" Font ที่ Footer Center Section ต้องเป็น ตัวหนา(Bold) เท่านั้น')
                else: print(f'Worksheet : "{sheet.title}" ขนาด Font ที่ Footer Center Section ต้องมีขนาด 18 เท่านั้น ขนาด Font ของคุณ : {center.size}')
            else: print(f'Worksheet : "{sheet.title}" Font ที่ Footer Center Section ต้องเป็น TH SarabunPSK เท่านั้น Font ของคุณ : {center.font}')
        else: print(f'Worksheet : "{sheet.title}" Font ที่ Footer Center Section ต้องเป็น TH SarabunPSK เท่านั้น Font ของคุณ : Calibri')
    else : print(f'Worksheet : "{sheet.title}" Footer Center section กำหนดให้พิมพ์ชื่อ-สกุลของตนเอง')
    
    if right.text.replace(" ","") == '&D&T' : 
        if right.font != None:
            if right.font.startswith("TH SarabunPSK"):
                if right.size == 18: 
                    if right.font.endswith("Bold") or right.font.endswith("ตัวหนา"): score += 1
                    else : print(f'Worksheet : "{sheet.title}" Font ที่ Footer Right Section ต้องเป็น ตัวหนา(Bold) เท่านั้น')
                else: print(f'Worksheet : "{sheet.title}" ขนาด Font ที่ Footer right Section ต้องมีขนาด 18 เท่านั้น ขนาด Font ของคุณ : {right.size}')
            else: print(f'Worksheet : "{sheet.title}" Font ที่ Footer Right Section ต้องเป็น TH SarabunPSK เท่านั้น Font ของคุณ : {right.font}')
        else: print(f'Worksheet : "{sheet.title}" Font ที่ Footer Right Section ต้องเป็น TH SarabunPSK เท่านั้น Font ของคุณ : Calibri')
    else : print(f'Worksheet : "{sheet.title}" Footer Right section กำหนดให้แทรกวันที่และเวลาอัตโนมัติจาก excel')
    return score












path = 'C:\\Users\\Aztaroz\\Documents\\test\\01_19_64100738 (version Thai).xlsx'
wb = openpyxl.load_workbook(path)
sheet = wb.worksheets[0]
score = 0

#Cover 
print("################# information of student #################")
cell_n = 1
while True:
    cellA = sheet['A{}'.format(cell_n)]
    cellB = sheet['B{}'.format(cell_n)]
    cell_n += 1
    #!print(cellA,cellA.value,cellB,cellB.value)
    print(cellA.value,cellB.value)
    #print(cell_n)
    if cell_n == 5:
        break
print("##########################################################")

# 1.
def check_worksheet(sheet,score): #sheet คือชื่อของ worksheet ที่ถูกส่งมาจาก name
    print("##### ข้อที่ 1 #####")
    sales_years = [62, 63, 64]
    index_sales_years = 0
    for a in range(len(wb.sheetnames)):
        sheet_index = wb.sheetnames.index(wb.sheetnames[a])
        if wb.sheetnames[a].upper() == 'COVER':
            score += 1
            #!print(f"ชื่อ worksheet : {wb.sheetnames[a]} index ที่ : {sheet_index} (ถูกต้อง)")
        elif wb.sheetnames[a].replace(" ","") == f'ยอดขายปี25{sales_years[index_sales_years]}':
            #!print(f"ชื่อ worksheet : {wb.sheetnames[a]} index ที่ : {sheet_index} (ถูกต้อง)")
            index_sales_years += 1 
            score += 1
        else:
            #print("รูปแบบที่ถูกต้อง คือ COVER" if sheet_index == 0 else (print(f"ยอดขายปี25{sales_years[index_sales_years]}"), index_sales_years + 1)[0])
            
            if sheet_index == 0 : print("รูปแบบที่ถูกต้อง คือ COVER")
            else : 
                print(f'รูปแบบที่ถูกต้อง คือ ยอดขายปี25{sales_years[index_sales_years]}')
                index_sales_years += 1
            
    return tabColor(sheet,score)

def tabColor(sheet,score):
    print("##### ข้อที่ 5 #####")
    sheet = wb.worksheets[2]
    if sheet.sheet_properties.tabColor.theme == 5: 
        score +=1
    else: 
        print("Tab Color Worksheet ที่ 2 ต้องเป็นสี Orange จาก Theme Color เท่านั้น")
    return checkA1Font(sheet,score)

def checkA1Font(sheet,score):
    print("##### ข้อที่ 6 #####")
    checkFont = ['A1','A2','B2','C3','D4','E5','F6','G7']
    for sheet in wb.worksheets:
        if sheet.title.upper() != "COVER":
            #!print(f"##########{sheet}##########")
            #!print(f"Score : {score}")
            for cellIndex in checkFont:
                font = sheet[cellIndex].font.name
                # Check A1
                if cellIndex == 'A1' and sheet[cellIndex].font.b == True :
                    score += 0.5
                elif cellIndex == 'A1' and sheet[cellIndex].font.b != True :
                    print(f"{sheet.title}: Font ใน Cell ที่ {cellIndex} ต้องเป็นตัวหนาเท่านั้น")

                # Check Other
                if sheet[cellIndex].font.name == 'TH SarabunPSK':
                    score += 0.5
                else: 
                    print(f"{sheet.title} : Font ใน Cell ที่ {cellIndex} ต้องเป็น TH SarabunPSK เท่านั้น Font ของคุณ : {font}")
                
                font = sheet[cellIndex].font.sz

                if sheet[cellIndex].font.sz == 20:
                    score += 0.5
                else: 
                    print(f"{sheet.title} : Font ใน Cell ที่ {cellIndex} ต้องขนาด 20 Point เท่านั้น Font Size ของคุณ : {font}")
                #!print(cellIndex,score)
    return checkRowHeight(sheet,score)

def checkRowHeight(sheet,score):
    print("##### ข้อที่ 7 #####")
    rowIndex = 1 # Row จะเริ่มที่ Row 1
    for sheet in wb.worksheets:
        rowIndex = 1
        for index in range(7):
            #print('height = ',sheet.row_dimensions[rowIndex].height, 'rowIndex = ',rowIndex)
            if sheet.row_dimensions[rowIndex].height != None and index <= 7 :
                rowHeight = 35 - sheet.row_dimensions[rowIndex].height
                if abs(rowHeight) >= 0 and abs(rowHeight) <= 0.6:
                    score += 1
                else: print(f"{sheet.title}/Row ที่ {rowIndex} : ความสูงของแถวต้องเป็น 35 เท่านั้น ความสูงที่คุณกำหนด {sheet.row_dimensions[rowIndex].height}")
                rowIndex += 1
            #print(score)
    return columnWidthCheck(sheet,score)

def columnWidthCheck(sheet,score):
    print("##### ข้อที่ 8 #####")
    for sheet in wb.worksheets:
        if notCover(sheet) == True:
            #print(sheet.title)
            num_col = 1
            for x in range(7):
                if notCover(sheet) == True:
                    # get_column_letter เป็นฟังก์ชันที่ใช้ในการแปลงตัวเลขเป็นตัวอักษรที่ใช้ในการกำหนดชื่อคอลัมน์
                    col_dim = sheet.column_dimensions[get_column_letter(num_col)] 
                    # .auto_size ใช้ในการเปลี่ยนขนาดความกว้างของ column โดยอัตโนมัติให้พอดีกับขนาดของข้อมูลที่อยู่ใน column นั้นๆ 
                    if col_dim.auto_size is True: # check if the column width is AutoFit
                        score += 1
                        #!print(f"Column {get_column_letter(num_col)} is AutoFit ('ถูกต้อง')")
                        
                    # ตรวจสอบว่ามีตัวแปร col_dim ใดๆ ในลิสต์ ['D', 'E', 'F', 'G'] ที่มีค่าเท่ากับ sheet.column_dimensions[col] ถ้าเป็นจริงอย่างน้อย 1 ตัวจะเข้าเงื่อนไข
                    # any() เพื่อตรวจสอบว่ามีค่าใดค่าหนึ่งใน iterable ที่เป็น True หรือไม่ 
                    # for col in ['D', 'E', 'F', 'G'] เป็น iterable object โดยใช้ list เป็น iterable object ในการวนลูปแต่ละรอบ
                    elif any(col_dim == sheet.column_dimensions[col] for col in ['D', 'E', 'F', 'G']):
                        #print(col_dim.width)
                        fix_col = 14.29 - col_dim.width
                        if abs(fix_col) <= 1.30:
                            score += 1
                            #!print(f"Column {get_column_letter(num_col)} width: {col_dim.width} ('ถูกต้อง')")
                        else:
                            print(f"Column {get_column_letter(num_col)} width: {col_dim.width} เกินค่าที่กำหนดไว้")
                    else:
                        print(f"Column {get_column_letter(num_col)} is not AutoFit ('ไม่ถูกต้อง')")
                    num_col += 1
    # print(f"Total score: {score}")
    return checkPageSetup(sheet,score)

def checkPageSetup(sheet,score):
    print("##### ข้อที่ 9 #####")
    for index in range(len(wb.sheetnames)):
        if index != 0:
            sheet = wb.worksheets[index]
            if sheet.page_setup.paperSize == 9: score += 1
            else: print(f"{sheet.title} : ขนาดกระดาษต้องเป็น A4 เท่านั้น")
            if sheet.page_setup.orientation == 'landscape': score += 1
            else: print(f"{sheet.title} : Orientation ต้องเป็น Landscape เท่านั้น")
                # Margin
                
        ###################################   Edit Score When merging   ##################################################################
            if marginTop(sheet) == True : score += 1
            if marginLeft(sheet) == True : score += 1
            if marginRight(sheet) == True : score += 1
            if marginBottom(sheet) == True : score += 1
                    
            if marginHeader(sheet) == True : score += 1
            if marginFooter(sheet) == True : score += 1

            if horizontalCenter(sheet) == True : score += 1
                
        ###################################   Header   ##################################################################
            if headerCheck(sheet) == True: score += 1

        ###################################   Footer   ##################################################################
            score + footerCheck(sheet,score)
    return score

score = check_worksheet(sheet,score)
print('final',score)