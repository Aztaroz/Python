import openpyxl
# path = 'C:\\Users\\Aztaroz\\Documents\\test\\01_19_64100738 (version Thai).xlsx'
# wb = openpyxl.load_workbook(path)
# sheet = wb.worksheets[0]

class Lab01:
    def __init__(self, path):
        self.wb = openpyxl.load_workbook(path)
        self.sheet = self.wb.worksheets[0]
        self.id = self.sheet.cell(3,2).value
        self.name = self.sheet.cell(1,2).value
        self.lastname = self.sheet.cell(2,2).value
        self.section = self.sheet.cell(4,2).value
        self.score = 0

        print(self.id,self.name,self.lastname, f'section {self.section}\nscore = {self.score}')

    def info(self):
        print(self.id,self.name,self.lastname, f'section {self.section}\nscore = {self.score}')

    def check(self,score):
        

a = Lab01('C:\\Users\\Aztaroz\\Documents\\test\\01_19_64100738 (version Thai).xlsx')
b = Lab01('C:\\Users\\Aztaroz\\Documents\\test\\01_19_64100738 (version Thai) - Copy.xlsx')